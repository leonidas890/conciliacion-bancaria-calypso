import streamlit as st
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import re
from io import BytesIO
import warnings
warnings.filterwarnings('ignore')

st.set_page_config(
    page_title="Conciliación Bancaria Automática",
    page_icon="📊",
    layout="wide"
)

# Inicializar session_state para evitar errores de DOM
if 'results' not in st.session_state:
    st.session_state['results'] = None
if 'sheet_names' not in st.session_state:
    st.session_state['sheet_names'] = None
if 'original_files_data' not in st.session_state:
    st.session_state['original_files_data'] = {}

# Estilos CSS personalizados - Diseño elegante y moderno
st.markdown("""
<style>
    /* Fondo general elegante con gradiente suave */
    .stApp {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 25%, #f093fb 50%, #4facfe 75%, #00f2fe 100%);
        background-size: 400% 400%;
        animation: gradientShift 15s ease infinite;
        min-height: 100vh;
    }
    
    @keyframes gradientShift {
        0% { background-position: 0% 50%; }
        50% { background-position: 100% 50%; }
        100% { background-position: 0% 50%; }
    }
    
    /* Header principal con logo CALYPSO - Fondo rojo como en la imagen */
    .calypso-header {
        background: linear-gradient(135deg, #DC143C 0%, #C41E3A 50%, #B22222 100%);
        padding: 3rem 2rem;
        border-radius: 20px;
        box-shadow: 0 8px 32px rgba(220, 20, 60, 0.4), inset 0 1px 0 rgba(255, 255, 255, 0.1);
        margin-bottom: 2.5rem;
        text-align: center;
        border: none;
        transition: transform 0.3s ease, box-shadow 0.3s ease;
        position: relative;
        overflow: hidden;
    }
    
    .calypso-header::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        bottom: 0;
        background: radial-gradient(ellipse at center top, rgba(255, 255, 255, 0.1) 0%, transparent 70%);
        pointer-events: none;
    }
    
    .calypso-header:hover {
        transform: translateY(-3px);
        box-shadow: 0 12px 40px rgba(220, 20, 60, 0.5), inset 0 1px 0 rgba(255, 255, 255, 0.15);
    }
    
    /* Contenedor para la imagen del logo */
    .logo-container {
        background: linear-gradient(135deg, #DC143C 0%, #C41E3A 50%, #B22222 100%);
        padding: 2rem;
        border-radius: 20px;
        box-shadow: 0 8px 32px rgba(220, 20, 60, 0.4), inset 0 1px 0 rgba(255, 255, 255, 0.1);
        margin-bottom: 2.5rem;
        text-align: center;
        position: relative;
        overflow: hidden;
    }
    
    .logo-container::before {
        content: '';
        position: absolute;
        top: 0;
        left: 0;
        right: 0;
        bottom: 0;
        background: radial-gradient(ellipse at center top, rgba(255, 255, 255, 0.1) 0%, transparent 70%);
        pointer-events: none;
    }
    
    /* Estilo para la imagen del logo */
    .stImage img {
        max-width: 100%;
        height: auto;
        border-radius: 10px;
    }
    
    .calypso-logo {
        font-size: 4.5rem;
        font-weight: 900;
        color: #FFFFFF;
        letter-spacing: 5px;
        margin: 0;
        text-shadow: 0 2px 4px rgba(0, 0, 0, 0.3);
        font-family: 'Arial Black', 'Helvetica Neue', 'Arial', sans-serif;
        text-transform: uppercase;
        position: relative;
        z-index: 1;
        font-variant: normal;
        /* Asegurar que se use Y (Y griega) y no I (I latina) */
        unicode-bidi: bidi-override;
    }
    
    .calypso-tagline {
        font-size: 1rem;
        color: #FFFFFF;
        margin-top: 1rem;
        letter-spacing: 3px;
        font-weight: 600;
        text-transform: uppercase;
        text-shadow: 0 1px 2px rgba(0, 0, 0, 0.2);
        position: relative;
        z-index: 1;
    }
    
    .calypso-trademark {
        font-size: 1.2rem;
        color: #FFFFFF;
        vertical-align: super;
        font-weight: 400;
        position: relative;
        z-index: 1;
    }
    
    .main-header {
        font-size: 2rem;
        font-weight: 700;
        background: linear-gradient(135deg, #ffffff 0%, #f0f0f0 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
        text-align: center;
        padding: 1.5rem 0;
        margin-top: 1rem;
        text-shadow: 0 2px 4px rgba(255, 255, 255, 0.3);
    }
    
    /* Tarjetas de estadísticas elegantes con glassmorphism */
    .stat-card {
        background: linear-gradient(135deg, rgba(255, 255, 255, 0.25) 0%, rgba(255, 255, 255, 0.15) 100%);
        backdrop-filter: blur(10px);
        padding: 2rem 1.5rem;
        border-radius: 20px;
        color: #ffffff;
        text-align: center;
        margin: 1rem 0;
        border: 1px solid rgba(255, 255, 255, 0.3);
        box-shadow: 0 8px 32px rgba(0, 0, 0, 0.1), inset 0 1px 0 rgba(255, 255, 255, 0.2);
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        position: relative;
        overflow: hidden;
    }
    
    .stat-card::before {
        content: '';
        position: absolute;
        top: 0;
        left: -100%;
        width: 100%;
        height: 100%;
        background: linear-gradient(90deg, transparent, rgba(255, 255, 255, 0.2), transparent);
        transition: left 0.5s;
    }
    
    .stat-card:hover::before {
        left: 100%;
    }
    
    .stat-card:hover {
        transform: translateY(-8px) scale(1.02);
        box-shadow: 0 12px 40px rgba(0, 0, 0, 0.2), inset 0 1px 0 rgba(255, 255, 255, 0.3);
        border-color: rgba(255, 255, 255, 0.5);
    }
    
    .stat-value {
        font-size: 2.5rem;
        font-weight: 800;
        color: #ffffff;
        text-shadow: 0 2px 8px rgba(0, 0, 0, 0.2);
        margin-bottom: 0.5rem;
    }
    
    .stat-label {
        font-size: 1rem;
        color: rgba(255, 255, 255, 0.9);
        margin-top: 0.5rem;
        font-weight: 500;
        text-transform: uppercase;
        letter-spacing: 1px;
    }
    
    .success {
        background: linear-gradient(135deg, rgba(46, 213, 115, 0.3) 0%, rgba(39, 174, 96, 0.3) 100%);
        border-left: 4px solid #2ed573;
    }
    
    .success .stat-value {
        color: #2ed573;
        text-shadow: 0 0 20px rgba(46, 213, 115, 0.5);
    }
    
    .warning {
        background: linear-gradient(135deg, rgba(255, 159, 67, 0.3) 0%, rgba(255, 107, 107, 0.3) 100%);
        border-left: 4px solid #ff9f43;
    }
    
    .warning .stat-value {
        color: #ff9f43;
        text-shadow: 0 0 20px rgba(255, 159, 67, 0.5);
    }
    
    .info {
        background: linear-gradient(135deg, rgba(52, 152, 219, 0.3) 0%, rgba(41, 128, 185, 0.3) 100%);
        border-left: 4px solid #3498db;
    }
    
    .info .stat-value {
        color: #3498db;
        text-shadow: 0 0 20px rgba(52, 152, 219, 0.5);
    }
    
    /* Sidebar elegante */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, rgba(102, 126, 234, 0.9) 0%, rgba(118, 75, 162, 0.9) 100%);
        backdrop-filter: blur(10px);
    }
    
    [data-testid="stSidebar"] [data-testid="stMarkdownContainer"] {
        color: #ffffff;
    }
    
    /* Botones elegantes con efectos */
    .stButton > button {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        color: white;
        border: none;
        border-radius: 12px;
        padding: 0.75rem 2rem;
        font-weight: 600;
        font-size: 1rem;
        transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
        box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);
        position: relative;
        overflow: hidden;
    }
    
    .stButton > button::before {
        content: '';
        position: absolute;
        top: 50%;
        left: 50%;
        width: 0;
        height: 0;
        border-radius: 50%;
        background: rgba(255, 255, 255, 0.3);
        transform: translate(-50%, -50%);
        transition: width 0.6s, height 0.6s;
    }
    
    .stButton > button:hover::before {
        width: 300px;
        height: 300px;
    }
    
    .stButton > button:hover {
        transform: translateY(-3px) scale(1.05);
        box-shadow: 0 8px 25px rgba(102, 126, 234, 0.6);
        background: linear-gradient(135deg, #764ba2 0%, #667eea 100%);
    }
    
    .stButton > button:active {
        transform: translateY(-1px) scale(1.02);
    }
    
    /* Inputs elegantes */
    .stFileUploader {
        background: rgba(255, 255, 255, 0.2);
        backdrop-filter: blur(10px);
        border-radius: 15px;
        padding: 1.5rem;
        border: 2px dashed rgba(255, 255, 255, 0.5);
        transition: all 0.3s ease;
    }
    
    .stFileUploader:hover {
        background: rgba(255, 255, 255, 0.3);
        border-color: rgba(255, 255, 255, 0.7);
    }
    
    /* Tablas elegantes */
    .dataframe {
        background: rgba(255, 255, 255, 0.15);
        backdrop-filter: blur(10px);
        border-radius: 15px;
        color: #ffffff;
        border: 1px solid rgba(255, 255, 255, 0.2);
    }
    
    /* Mensajes elegantes */
    .stSuccess {
        background: linear-gradient(135deg, rgba(46, 213, 115, 0.9) 0%, rgba(39, 174, 96, 0.9) 100%);
        border-left: 4px solid #2ed573;
        color: #ffffff;
        border-radius: 10px;
        box-shadow: 0 4px 15px rgba(46, 213, 115, 0.3);
    }
    
    .stInfo {
        background: linear-gradient(135deg, rgba(52, 152, 219, 0.9) 0%, rgba(41, 128, 185, 0.9) 100%);
        border-left: 4px solid #3498db;
        color: #ffffff;
        border-radius: 10px;
        box-shadow: 0 4px 15px rgba(52, 152, 219, 0.3);
    }
    
    .stError {
        background: linear-gradient(135deg, rgba(231, 76, 60, 0.9) 0%, rgba(192, 57, 43, 0.9) 100%);
        border-left: 4px solid #e74c3c;
        color: #ffffff;
        border-radius: 10px;
        box-shadow: 0 4px 15px rgba(231, 76, 60, 0.3);
    }
    
    .stWarning {
        background: linear-gradient(135deg, rgba(243, 156, 18, 0.9) 0%, rgba(230, 126, 34, 0.9) 100%);
        border-left: 4px solid #f39c12;
        color: #ffffff;
        border-radius: 10px;
        box-shadow: 0 4px 15px rgba(243, 156, 18, 0.3);
    }
    
    /* Texto general elegante */
    .stMarkdown, .stText {
        color: #ffffff;
        text-shadow: 0 1px 3px rgba(0, 0, 0, 0.2);
    }
    
    /* Headers elegantes */
    h1, h2, h3, h4, h5, h6 {
        color: #ffffff !important;
        text-shadow: 0 2px 4px rgba(0, 0, 0, 0.2);
        font-weight: 700;
    }
    
    /* Scrollbar elegante */
    ::-webkit-scrollbar {
        width: 10px;
    }
    
    ::-webkit-scrollbar-track {
        background: rgba(255, 255, 255, 0.1);
        border-radius: 10px;
    }
    
    ::-webkit-scrollbar-thumb {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        border-radius: 10px;
    }
    
    ::-webkit-scrollbar-thumb:hover {
        background: linear-gradient(135deg, #764ba2 0%, #667eea 100%);
    }
    
    /* Efectos de entrada suaves */
    @keyframes fadeInUp {
        from {
            opacity: 0;
            transform: translateY(20px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    .stat-card {
        animation: fadeInUp 0.6s ease-out;
    }
</style>
""", unsafe_allow_html=True)

def normalize_date(date_value):
    """Normaliza fechas a formato YYYY-MM-DD - Versión mejorada para coincidencias exactas"""
    if pd.isna(date_value) or date_value == '':
        return ''
    
    # Si es datetime o Timestamp, convertir directamente
    if isinstance(date_value, (pd.Timestamp, datetime)):
        try:
            return date_value.strftime('%Y-%m-%d')
        except:
            return ''
    
    # Si es string, intentar parsear
    if isinstance(date_value, str):
        date_str = str(date_value).strip()
        
        # Si ya está en formato YYYY-MM-DD, retornar directamente
        if re.match(r'^\d{4}-\d{2}-\d{2}$', date_str):
            return date_str
        
        # Formatos comunes con regex más estricto
        patterns = [
            # DD/MM/YYYY o D/M/YYYY
            (r'^(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})$', lambda m: f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"),
            # YYYY-MM-DD o YYYY/MM/DD
            (r'^(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})$', lambda m: f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"),
            # DD.MM.YYYY
            (r'^(\d{1,2})\.(\d{1,2})\.(\d{4})$', lambda m: f"{m.group(3)}-{m.group(2).zfill(2)}-{m.group(1).zfill(2)}"),
            # YYYY.MM.DD
            (r'^(\d{4})\.(\d{1,2})\.(\d{1,2})$', lambda m: f"{m.group(1)}-{m.group(2).zfill(2)}-{m.group(3).zfill(2)}"),
        ]
        
        for pattern, formatter in patterns:
            match = re.match(pattern, date_str)
            if match:
                try:
                    formatted = formatter(match)
                    # Validar que la fecha sea válida
                    year, month, day = map(int, formatted.split('-'))
                    if 1900 <= year <= 2100 and 1 <= month <= 12 and 1 <= day <= 31:
                        return formatted
                except:
                    continue
        
        # Intentar parsear con pandas (más flexible)
        try:
            parsed = pd.to_datetime(date_str, dayfirst=True, errors='coerce')
            if pd.notna(parsed):
                return parsed.strftime('%Y-%m-%d')
        except:
            pass
        
        # Si no se pudo parsear, retornar string vacío para evitar falsos positivos
        return ''
    
    # Si es número (fecha serial de Excel), convertir
    if isinstance(date_value, (int, float)):
        try:
            # Intentar convertir desde fecha serial de Excel
            if date_value > 1 and date_value < 1000000:  # Rango razonable para fechas Excel
                excel_date = pd.Timestamp('1900-01-01') + pd.Timedelta(days=int(date_value) - 2)
                return excel_date.strftime('%Y-%m-%d')
        except:
            pass
    
    return ''

def extract_pv(text):
    """Extrae código PV de un texto - Versión mejorada para coincidencias exactas, incluyendo LOG"""
    if pd.isna(text) or text == '':
        return ''
    
    text_str = str(text).upper().strip()
    
    # Si está vacío después de limpiar, retornar vacío
    if not text_str:
        return ''
    
    # Caso 1: LOG seguido de números (LOG81, LOG081, LOG 81, etc.)
    log_match = re.search(r'LOG\s*0*(\d+)', text_str)
    if log_match:
        num = log_match.group(1)
        # Normalizar LOG a PV (LOG81 -> PV081)
        return f"PV{num.zfill(3)}"
    
    # Caso 2: Ya está en formato PV### (con o sin ceros a la izquierda)
    pv_match = re.search(r'PV\s*0*(\d+)', text_str)
    if pv_match:
        num = pv_match.group(1)
        # Normalizar a 3 dígitos (rellenar con ceros a la izquierda)
        return f"PV{num.zfill(3)}"
    
    # Caso 3: Es solo un número (sin PV ni LOG) - como "81"
    num_match = re.match(r'^0*(\d+)$', text_str)
    if num_match:
        num = num_match.group(1)
        return f"PV{num.zfill(3)}"
    
    # Caso 4: Buscar número significativo (2 o más dígitos) en el texto
    # Priorizar números más largos al final
    numbers = re.findall(r'(\d{2,})', text_str)
    if numbers:
        # Tomar el último número encontrado (más probable que sea el PV)
        num = numbers[-1]
        return f"PV{num.zfill(3)}"
    
    # Caso 5: Buscar cualquier número (1 o más dígitos)
    any_num = re.search(r'(\d+)', text_str)
    if any_num:
        num = any_num.group(1)
        return f"PV{num.zfill(3)}"
    
    # Caso 6: Limpiar y buscar cualquier alfanumérico que pueda ser PV
    cleaned = re.sub(r'[^A-Z0-9]', '', text_str)
    if cleaned and len(cleaned) >= 2:
        # Si tiene al menos 2 caracteres, intentar extraer número
        num_in_cleaned = re.search(r'(\d+)', cleaned)
        if num_in_cleaned:
            num = num_in_cleaned.group(1)
            return f"PV{num.zfill(3)}"
    
    return ''

def normalize_amount(value):
    """Normaliza montos a número"""
    if pd.isna(value):
        return 0.0
    
    if isinstance(value, (int, float)):
        return abs(float(value))
    
    if isinstance(value, str):
        # Limpiar string
        cleaned = re.sub(r'[$€£¥₱₹¢\s]', '', value)
        
        # Detectar formato
        if ',' in cleaned and '.' in cleaned:
            # Determinar cuál es decimal
            last_comma = cleaned.rfind(',')
            last_dot = cleaned.rfind('.')
            if last_comma > last_dot:
                # Formato europeo: 1.234,56
                cleaned = cleaned.replace('.', '').replace(',', '.')
            else:
                # Formato americano: 1,234.56
                cleaned = cleaned.replace(',', '')
        elif ',' in cleaned:
            # Solo coma, puede ser decimal
            if re.search(r',\d{1,2}$', cleaned):
                cleaned = cleaned.replace(',', '.')
            else:
                cleaned = cleaned.replace(',', '')
        else:
            # Solo punto o sin separadores
            pass
        
        try:
            return abs(float(re.sub(r'[^0-9.-]', '', cleaned)))
        except:
            return 0.0
    
    return 0.0

def detect_columns(df):
    """Detecta automáticamente las columnas de fecha, monto y referencia - Versión mejorada"""
    if df.empty or len(df.columns) == 0:
        return None, None, None, None
    
    # Crear diccionario más flexible de búsqueda
    columns_lower = {}
    for col in df.columns:
        # Normalizar nombre de columna
        normalized = str(col).lower().strip()
        # Remover espacios, guiones, puntos, etc.
        normalized_clean = re.sub(r'[_\s\-\.]', '', normalized)
        columns_lower[normalized_clean] = col
        # También guardar versión con espacios
        columns_lower[normalized] = col
    
    fecha_col = None
    monto_col = None
    referencia_col = None
    descripcion_col = None
    
    # Buscar fecha - MÁS VARIACIONES
    fecha_patterns = [
        'fecha', 'date', 'fec', 'dia', 'day', 'fechapago', 'fechaoperacion',
        'fechacontabilizacion', 'fechavencimiento', 'foperacion', 'fpago',
        'fechadecontabilizacion', 'fechadevencimiento'
    ]
    for pattern in fecha_patterns:
        if pattern in columns_lower:
            fecha_col = columns_lower[pattern]
            break
    
    # Si no se encontró, buscar por contenido (columnas que parecen fechas)
    if not fecha_col:
        for col in df.columns:
            # Verificar si la columna contiene fechas
            sample_values = df[col].dropna().head(10)
            date_count = 0
            for val in sample_values:
                try:
                    pd.to_datetime(val)
                    date_count += 1
                except:
                    pass
            if date_count >= 3:  # Si al menos 3 valores son fechas
                fecha_col = col
                break
    
    # Buscar monto - MÁS VARIACIONES
    monto_patterns = [
        'monto', 'amount', 'importe', 'valor', 'total', 'cantidad', 'pago', 
        'abono', 'cargo', 'credito', 'debito', 'value', 'cargoabono',
        'cargoabonoml', 'saldo', 'suma'
    ]
    for pattern in monto_patterns:
        if pattern in columns_lower:
            monto_col = columns_lower[pattern]
            break
    
    # Si no se encontró, buscar columnas numéricas
    if not monto_col:
        for col in df.columns:
            # Verificar si la columna es numérica
            if df[col].dtype in ['int64', 'float64']:
                # Verificar que tenga valores significativos
                if df[col].abs().sum() > 0:
                    monto_col = col
                    break
            else:
                # Intentar convertir a numérico
                try:
                    numeric_vals = pd.to_numeric(df[col].astype(str).str.replace(r'[^\d.-]', '', regex=True), errors='coerce')
                    if numeric_vals.notna().sum() > len(df) * 0.5:  # Si más del 50% son numéricos
                        monto_col = col
                        break
                except:
                    pass
    
    # Buscar referencia/PV - MÁS VARIACIONES
    referencia_patterns = [
        'referencia', 'ref', 'reference', 'pv', 'puntoventa', 'puntoventa',
        'numero', 'num', 'codigopv', 'codpv', 'nro', 'no', 'voucher',
        'comprobante', 'ticket', 'folio', 'numerooperacion', 'nrooperacion',
        'comentarios', 'codpv', 'codigopuntoventa'
    ]
    for pattern in referencia_patterns:
        if pattern in columns_lower:
            referencia_col = columns_lower[pattern]
            break
    
    # Buscar descripción
    descripcion_patterns = [
        'descripcion', 'description', 'desc', 'concepto', 'detalle', 
        'observacion', 'nota', 'comentario', 'memo', 'nombredelacuentadecontrapartida',
        'nombrepv', 'glosa'
    ]
    for pattern in descripcion_patterns:
        if pattern in columns_lower:
            descripcion_col = columns_lower[pattern]
            break
    
    return fecha_col, monto_col, referencia_col, descripcion_col

def process_excel_file(file, sheet_name=None, fecha_col=None, monto_col=None, referencia_col=None, descripcion_col=None):
    """Procesa un archivo Excel y retorna datos normalizados"""
    try:
        if sheet_name:
            df = pd.read_excel(file, sheet_name=sheet_name, engine='openpyxl')
        else:
            df = pd.read_excel(file, engine='openpyxl')
        
        # Si no se proporcionaron columnas, detectarlas automáticamente
        if not fecha_col or not monto_col:
            detected_fecha, detected_monto, detected_ref, detected_desc = detect_columns(df)
            
            if not fecha_col:
                fecha_col = detected_fecha
            if not monto_col:
                monto_col = detected_monto
            if not referencia_col:
                referencia_col = detected_ref
            if not descripcion_col:
                descripcion_col = detected_desc
        
        if not fecha_col or not monto_col:
            # Preparar información de columnas disponibles para el error
            available_cols = ", ".join(df.columns.astype(str).tolist())
            return None, f"No se pudieron detectar las columnas de fecha y monto. Columnas disponibles: {available_cols}"
        
        # Normalizar datos - Guardar índice real del DataFrame
        processed_data = []
        for idx, row in df.iterrows():
            fecha = normalize_date(row.get(fecha_col, ''))
            monto = normalize_amount(row.get(monto_col, 0))
            referencia = extract_pv(row.get(referencia_col, '')) if referencia_col else ''
            descripcion = str(row.get(descripcion_col, '')) if descripcion_col else ''
            
            if fecha and monto > 0:
                # Guardar el índice real del DataFrame (idx puede ser cualquier número)
                # El número de fila en Excel será idx + 2 (1 para encabezado + 1 porque Excel empieza en 1)
                processed_data.append({
                    'fecha': fecha,
                    'monto': round(monto, 2),
                    'referencia': referencia,
                    'descripcion': descripcion,
                    '_original': row.to_dict(),
                    '_excel_row': int(idx) + 2  # +2 porque Excel: fila 1 = encabezado, fila 2+ = datos
                })
        
        return processed_data, None
    except Exception as e:
        return None, f"Error al procesar archivo: {str(e)}"

def reconcile_data(data1, data2, name1="Hoja 1", name2="Hoja 2"):
    """Reconcilia dos conjuntos de datos - Busca automáticamente valores iguales por Fecha y PV"""
    results = []
    matched1 = set()
    matched2 = set()
    
    # Normalizar y validar datos antes de cruzar - NORMALIZACIÓN MEJORADA
    normalized_data1 = []
    normalized_data2 = []
    
    def normalize_for_matching(fecha, monto, referencia):
        """Normaliza valores para comparación exacta"""
        # Normalizar fecha: asegurar formato YYYY-MM-DD consistente
        fecha_norm = ''
        if fecha:
            fecha_str = str(fecha).strip()
            # Si ya está en formato YYYY-MM-DD, usar directamente
            if re.match(r'^\d{4}-\d{2}-\d{2}$', fecha_str):
                fecha_norm = fecha_str
            else:
                # Normalizar usando la función normalize_date
                fecha_norm = normalize_date(fecha)
        
        # Normalizar monto: redondear a 2 decimales y usar tolerancia
        monto_norm = 0.0
        if monto:
            try:
                monto_float = float(monto)
                monto_norm = round(abs(monto_float), 2)
            except:
                monto_norm = 0.0
        
        # Normalizar referencia/PV: extraer y normalizar PV
        ref_norm = ''
        if referencia:
            ref_str = str(referencia).strip().upper()
            # Si ya está en formato PV###, usar directamente
            if re.match(r'^PV\d+$', ref_str):
                # Extraer número y normalizar a PV### (3 dígitos)
                num_match = re.search(r'(\d+)', ref_str)
                if num_match:
                    ref_norm = f"PV{num_match.group(1).zfill(3)}"
            else:
                # Extraer PV usando extract_pv
                ref_norm = extract_pv(referencia)
                # Asegurar formato PV### si se encontró algo
                if ref_norm and not ref_norm.startswith('PV'):
                    num_match = re.search(r'(\d+)', ref_norm)
                    if num_match:
                        ref_norm = f"PV{num_match.group(1).zfill(3)}"
        
        return fecha_norm, monto_norm, ref_norm
    
    for idx, row in enumerate(data1):
        if row.get('fecha') and row.get('monto', 0) > 0:
            fecha_norm, monto_norm, ref_norm = normalize_for_matching(
                row.get('fecha', ''),
                row.get('monto', 0),
                row.get('referencia', '')
            )
            if fecha_norm:  # Solo agregar si tiene fecha válida
                # Usar _excel_row si existe (número de fila real en Excel), sino calcular desde índice
                excel_row = row.get('_excel_row')
                if excel_row is None:
                    # Si no tiene _excel_row, usar el índice + 2 (asumiendo que data1 viene de un DataFrame)
                    excel_row = idx + 2
                
                normalized_data1.append({
                    'fecha': fecha_norm,
                    'monto': monto_norm,
                    'referencia': ref_norm,
                    'descripcion': str(row.get('descripcion', '')),
                    '_original': row.get('_original', {}),
                    '_index_original': idx,
                    '_excel_row': excel_row  # Número de fila real en Excel
                })
    
    for idx, row in enumerate(data2):
        if row.get('fecha') and row.get('monto', 0) > 0:
            fecha_norm, monto_norm, ref_norm = normalize_for_matching(
                row.get('fecha', ''),
                row.get('monto', 0),
                row.get('referencia', '')
            )
            if fecha_norm:  # Solo agregar si tiene fecha válida
                # Usar _excel_row si existe (número de fila real en Excel), sino calcular desde índice
                excel_row = row.get('_excel_row')
                if excel_row is None:
                    # Si no tiene _excel_row, usar el índice + 2 (asumiendo que data2 viene de un DataFrame)
                    excel_row = idx + 2
                
                normalized_data2.append({
                    'fecha': fecha_norm,
                    'monto': monto_norm,
                    'referencia': ref_norm,
                    'descripcion': str(row.get('descripcion', '')),
                    '_original': row.get('_original', {}),
                    '_index_original': idx,
                    '_excel_row': excel_row  # Número de fila real en Excel
                })
    
    # Crear índices para búsqueda rápida en data2
    # PRIORIDAD 1: Fecha + PV + Monto (coincidencia EXACTA - máxima prioridad)
    data2_by_exact = {}  # fecha|pv|monto -> [indices]
    # PRIORIDAD 2: Fecha + PV (sin importar monto) - LO MÁS IMPORTANTE
    data2_by_date_pv = {}  # fecha|pv -> [indices]
    # PRIORIDAD 3: Fecha + Monto (sin PV)
    data2_by_date_amount = {}  # fecha|monto -> [indices]
    
    for idx, row2 in enumerate(normalized_data2):
        fecha = row2['fecha']
        pv = row2['referencia']
        monto = row2['monto']
        # Usar formato consistente para monto (siempre 2 decimales, sin ceros innecesarios)
        monto_str = f"{monto:.2f}".rstrip('0').rstrip('.')
        if not monto_str or monto_str == '.':
            monto_str = "0.00"
        
        # PRIORIDAD 1: Clave exacta: fecha + PV + monto (MÁXIMA PRIORIDAD)
        if fecha and pv:
            exact_key = f"{fecha}|{pv}|{monto_str}"
            if exact_key not in data2_by_exact:
                data2_by_exact[exact_key] = []
            data2_by_exact[exact_key].append(idx)
        
        # PRIORIDAD 2: Clave fecha + PV (sin importar monto) - LO MÁS IMPORTANTE
        if fecha and pv:
            date_pv_key = f"{fecha}|{pv}"
            if date_pv_key not in data2_by_date_pv:
                data2_by_date_pv[date_pv_key] = []
            data2_by_date_pv[date_pv_key].append(idx)
        
        # PRIORIDAD 3: Clave fecha + monto (sin importar PV)
        if fecha:
            date_amount_key = f"{fecha}|{monto_str}"
            if date_amount_key not in data2_by_date_amount:
                data2_by_date_amount[date_amount_key] = []
            data2_by_date_amount[date_amount_key].append(idx)
    
    # PASO 1: PRIORIDAD MÁXIMA - Coincidencias EXACTAS por FECHA + PV + MONTO
    # Primero buscar coincidencias exactas (todos los campos iguales)
    for idx1, row1 in enumerate(normalized_data1):
        if idx1 in matched1:
            continue
        
        fecha = row1['fecha']
        pv = row1['referencia']
        monto = row1['monto']
        monto_str = f"{monto:.2f}".rstrip('0').rstrip('.')
        if not monto_str or monto_str == '.':
            monto_str = "0.00"
        
        if not fecha:
            continue
        
        # PRIORIDAD 1: Coincidencia EXACTA (fecha + PV + monto)
        if fecha and pv:
            exact_key = f"{fecha}|{pv}|{monto_str}"
            candidates = data2_by_exact.get(exact_key, [])
            
            for idx2 in candidates:
                if idx2 in matched2:
                    continue
                
                row2 = normalized_data2[idx2]
                # Verificación adicional: asegurar que todos los campos coincidan exactamente
                if (row2['fecha'] == fecha and 
                    row2['referencia'] == pv and 
                    abs(row2['monto'] - monto) < 0.005):  # Tolerancia muy pequeña para montos
                    # Usar _excel_row para obtener el número de fila real en Excel
                    fila_banco = row1.get('_excel_row', idx1 + 2)
                    fila_interno = row2.get('_excel_row', idx2 + 2)
                    results.append({
                        'banco': {**row1, 'origen': name1},
                        'interno': {**row2, 'origen': name2},
                        'estado': 'Conciliado',
                        'origen': 'exacto_fecha_pv_monto',
                        'fila_banco': fila_banco,
                        'fila_interno': fila_interno
                    })
                    matched1.add(idx1)
                    matched2.add(idx2)
                    break
            
            if idx1 in matched1:
                continue  # Ya se encontró coincidencia exacta, pasar a siguiente
        
        # PRIORIDAD 2: Coincidencia por FECHA + PV (sin importar monto)
        if fecha and pv:
            date_pv_key = f"{fecha}|{pv}"
            candidates = data2_by_date_pv.get(date_pv_key, [])
            
            for idx2 in candidates:
                if idx2 in matched2:
                    continue
                
                row2 = normalized_data2[idx2]
                # Verificar que fecha y PV coincidan exactamente
                if row2['fecha'] == fecha and row2['referencia'] == pv:
                    # Usar _excel_row para obtener el número de fila real en Excel
                    fila_banco = row1.get('_excel_row', idx1 + 2)
                    fila_interno = row2.get('_excel_row', idx2 + 2)
                    results.append({
                        'banco': {**row1, 'origen': name1},
                        'interno': {**row2, 'origen': name2},
                        'estado': 'Conciliado',
                        'origen': 'fecha_pv',
                        'fila_banco': fila_banco,
                        'fila_interno': fila_interno
                    })
                    matched1.add(idx1)
                    matched2.add(idx2)
                    break
    
    # PASO 2: Coincidencias por fecha + monto (sin PV) - Solo si no se encontró por fecha+PV
    for idx1, row1 in enumerate(normalized_data1):
        if idx1 in matched1:
            continue
        
        fecha = row1['fecha']
        monto = row1['monto']
        monto_str = f"{monto:.2f}".rstrip('0').rstrip('.')
        if not monto_str or monto_str == '.':
            monto_str = "0.00"
        
        if not fecha:
            continue
        
        date_amount_key = f"{fecha}|{monto_str}"
        candidates = data2_by_date_amount.get(date_amount_key, [])
        
        for idx2 in candidates:
            if idx2 in matched2:
                continue
            
            row2 = normalized_data2[idx2]
            # Verificar que fecha y monto coincidan exactamente (con tolerancia pequeña)
            if row2['fecha'] == fecha and abs(row2['monto'] - monto) < 0.005:
                # Usar _excel_row para obtener el número de fila real en Excel
                fila_banco = row1.get('_excel_row', idx1 + 2)
                fila_interno = row2.get('_excel_row', idx2 + 2)
                results.append({
                    'banco': {**row1, 'origen': name1},
                    'interno': {**row2, 'origen': name2},
                    'estado': 'Conciliado',
                    'origen': 'fecha_monto',
                    'fila_banco': fila_banco,
                    'fila_interno': fila_interno
                })
                matched1.add(idx1)
                matched2.add(idx2)
                break
    
    # PASO 3: No conciliados de hoja 1
    for idx1, row1 in enumerate(normalized_data1):
        if idx1 not in matched1:
            # Usar _excel_row para obtener el número de fila real en Excel
            fila_banco = row1.get('_excel_row', idx1 + 2)
            results.append({
                'banco': {**row1, 'origen': name1},
                'interno': None,
                'estado': 'No conciliado',
                'origen': name1,
                'fila_banco': fila_banco,
                'fila_interno': None
            })
    
    # PASO 4: No conciliados de hoja 2
    for idx2, row2 in enumerate(normalized_data2):
        if idx2 not in matched2:
            # Usar _excel_row para obtener el número de fila real en Excel
            fila_interno = row2.get('_excel_row', idx2 + 2)
            results.append({
                'banco': {'fecha': '', 'monto': 0, 'referencia': '', 'descripcion': '', 'origen': ''},
                'interno': {**row2, 'origen': name2},
                'estado': 'No conciliado',
                'origen': name2,
                'fila_banco': None,
                'fila_interno': fila_interno
            })
    
    return results

def main():
    # Header con logo CALYPSO usando la imagen real
    import os
    
    # Intentar cargar la imagen del logo
    logo_paths = [
        'static/calypso-logo.png',
        'src/assets/calypso-logo.png',
        'calypso-logo.png'
    ]
    
    logo_path = None
    for path in logo_paths:
        if os.path.exists(path):
            logo_path = path
            break
    
    # Mostrar logo con imagen SVG en alta resolución desde URL
    # Usar el SVG directamente para evitar pixelación
    st.markdown("""
    <div style="background: linear-gradient(135deg, #DC143C 0%, #C41E3A 50%, #B22222 100%);
                padding: 2.5rem 2rem;
                border-radius: 20px;
                box-shadow: 0 8px 32px rgba(220, 20, 60, 0.4), inset 0 1px 0 rgba(255, 255, 255, 0.1);
                margin-bottom: 2.5rem;
                text-align: center;
                position: relative;
                overflow: hidden;">
        <div style="display: flex; justify-content: center; align-items: center;">
            <img src="https://www.tiendascalypso.com/arquivos/calypso-logo-2023-HD.svg?v=638138836791530000" 
                 alt="CALYPSO Logo" 
                 style="max-width: 100%; height: auto; width: 70%; filter: brightness(0) invert(1);">
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    st.markdown('<div class="main-header">📊 Conciliación Bancaria Automática</div>', unsafe_allow_html=True)
    
    # Sidebar para selección de modo
    st.sidebar.title("Configuración")
    mode = st.sidebar.radio(
        "Modo de operación",
        ["Un archivo (múltiples hojas)", "Dos archivos separados"]
    )
    
    if mode == "Un archivo (múltiples hojas)":
        st.header("📁 Cargar archivo Excel con dos hojas")
        uploaded_file = st.file_uploader(
            "Selecciona un archivo Excel con al menos 2 hojas",
            type=['xlsx', 'xls'],
            help="La primera hoja se cruzará contra la segunda"
        )
        
        if uploaded_file:
            try:
                # Leer archivo Excel y obtener nombres de hojas
                uploaded_file.seek(0)  # Asegurar que el archivo esté al inicio
                excel_file = pd.ExcelFile(uploaded_file, engine='openpyxl')
                sheet_names = excel_file.sheet_names
                
                if len(sheet_names) < 2:
                    st.error(f"❌ El archivo solo tiene {len(sheet_names)} hoja(s). Se necesitan al menos 2 hojas para realizar la conciliación.")
                    st.info("💡 Asegúrate de que tu archivo Excel tenga al menos 2 hojas con datos.")
                else:
                    st.success(f"✅ Archivo cargado: **{uploaded_file.name}**")
                    st.info(f"📑 **Hojas encontradas ({len(sheet_names)}):** {', '.join(sheet_names)}")
                    st.info(f"🔄 Se conciliará: **{sheet_names[0]}** ↔ **{sheet_names[1]}**")
                    
                    # Mostrar preview de columnas de cada hoja
                    try:
                        uploaded_file.seek(0)
                        df_preview1 = pd.read_excel(uploaded_file, sheet_name=sheet_names[0], engine='openpyxl', nrows=3)
                        if not df_preview1.empty:
                            st.info(f"📋 **Columnas en '{sheet_names[0]}':** {', '.join(df_preview1.columns.astype(str).tolist()[:10])}" + 
                                   (f" ... (+{len(df_preview1.columns) - 10} más)" if len(df_preview1.columns) > 10 else ""))
                        
                        uploaded_file.seek(0)
                        df_preview2 = pd.read_excel(uploaded_file, sheet_name=sheet_names[1], engine='openpyxl', nrows=3)
                        if not df_preview2.empty:
                            st.info(f"📋 **Columnas en '{sheet_names[1]}':** {', '.join(df_preview2.columns.astype(str).tolist()[:10])}" + 
                                   (f" ... (+{len(df_preview2.columns) - 10} más)" if len(df_preview2.columns) > 10 else ""))
                        uploaded_file.seek(0)
                    except Exception as preview_error:
                        st.warning(f"⚠️ No se pudo mostrar el preview de columnas: {str(preview_error)}")
                        uploaded_file.seek(0)
                    
                    if st.button("🔄 Procesar y Conciliar", type="primary"):
                        with st.spinner("Procesando archivo..."):
                            try:
                                # Procesar primera hoja
                                uploaded_file.seek(0)  # Asegurar que el archivo esté al inicio
                                data1, error1 = process_excel_file(uploaded_file, sheet_names[0])
                                if error1:
                                    st.error(f"Error en hoja 1 ({sheet_names[0]}): {error1}")
                                    if "Columnas disponibles" in error1:
                                        st.warning("💡 Asegúrate de que las columnas tengan nombres que contengan 'fecha' o 'date' para fechas, y 'monto', 'amount' o 'valor' para montos.")
                                    return
                                
                                # Procesar segunda hoja
                                uploaded_file.seek(0)  # Resetear archivo
                                data2, error2 = process_excel_file(uploaded_file, sheet_names[1])
                                if error2:
                                    st.error(f"Error en hoja 2 ({sheet_names[1]}): {error2}")
                                    if "Columnas disponibles" in error2:
                                        st.warning("💡 Asegúrate de que las columnas tengan nombres que contengan 'fecha' o 'date' para fechas, y 'monto', 'amount' o 'valor' para montos.")
                                    return
                                
                                if not data1 or len(data1) == 0:
                                    st.error(f"La hoja '{sheet_names[0]}' está vacía o no contiene datos válidos")
                                    return
                                
                                if not data2 or len(data2) == 0:
                                    st.error(f"La hoja '{sheet_names[1]}' está vacía o no contiene datos válidos")
                                    return
                                
                                # Guardar datos originales para preservar formato en descarga
                                try:
                                    uploaded_file.seek(0)
                                    df1_original = pd.read_excel(uploaded_file, sheet_name=sheet_names[0], engine='openpyxl')
                                    uploaded_file.seek(0)
                                    df2_original = pd.read_excel(uploaded_file, sheet_name=sheet_names[1], engine='openpyxl')
                                except Exception as e:
                                    st.warning(f"Advertencia: No se pudieron cargar los datos originales para preservar formato: {str(e)}")
                                    # Crear DataFrames vacíos como fallback
                                    df1_original = pd.DataFrame()
                                    df2_original = pd.DataFrame()
                                
                                # Reconciliar
                                results = reconcile_data(data1, data2, sheet_names[0], sheet_names[1])
                                
                                if not results or len(results) == 0:
                                    st.warning("⚠️ No se generaron resultados de conciliación. Verifica que los datos sean válidos.")
                                    return
                                
                                # Calcular estadísticas
                                conciliados = [r for r in results if r['estado'] == 'Conciliado']
                                exactos = [r for r in conciliados if r['origen'] == 'exacto_fecha_pv_monto']
                                fecha_pv = [r for r in conciliados if r['origen'] == 'fecha_pv']
                                fecha_monto = [r for r in conciliados if r['origen'] == 'fecha_monto']
                                
                                # Guardar resultados en session state
                                st.session_state['results'] = results
                                st.session_state['sheet_names'] = sheet_names
                                st.session_state['original_files_data'] = {
                                    'original_df1': df1_original,
                                    'original_df2': df2_original
                                }
                                
                                st.success(f"✅ Conciliación completada!")
                                st.info(f"""
                                **Resultados del cruce:**
                                - Total registros procesados: **{sheet_names[0]}**: {len(data1)}, **{sheet_names[1]}**: {len(data2)}
                                - **Conciliados: {len(conciliados)}**
                                  - ✅ Exactos (Fecha+PV+Monto): {len(exactos)}
                                  - 🔄 Por Fecha+PV: {len(fecha_pv)}
                                  - 📅 Por Fecha+Monto: {len(fecha_monto)}
                                - ❌ No conciliados: {len(results) - len(conciliados)}
                                """)
                                # Evitar st.rerun() que puede causar errores de DOM
                                # Los resultados se mostrarán automáticamente en la siguiente sección
                            except Exception as e:
                                st.error(f"❌ Error al procesar el archivo: {str(e)}")
                                import traceback
                                st.exception(e)
            except Exception as e:
                st.error(f"Error al leer el archivo: {str(e)}")
    
    else:  # Dos archivos separados
        st.header("📁 Cargar dos archivos Excel")
        col1, col2 = st.columns(2)
        
        with col1:
            banco_file = st.file_uploader(
                "Archivo de Pagos Bancarios",
                type=['xlsx', 'xls', 'csv'],
                key="banco"
            )
        
        with col2:
            interno_file = st.file_uploader(
                "Archivo de Pagos Internos",
                type=['xlsx', 'xls', 'csv'],
                key="interno"
            )
        
        if banco_file and interno_file:
            if st.button("🔄 Procesar y Conciliar", type="primary"):
                with st.spinner("Procesando archivos..."):
                    # Procesar archivo banco
                    data1, error1 = process_excel_file(banco_file)
                    if error1:
                        st.error(f"Error en archivo bancario: {error1}")
                        return
                    
                    # Procesar archivo interno
                    data2, error2 = process_excel_file(interno_file)
                    if error2:
                        st.error(f"Error en archivo interno: {error2}")
                        return
                    
                    if not data1 or not data2:
                        st.error("Uno o ambos archivos están vacíos")
                        return
                    
                    # Guardar datos originales para preservar formato
                    banco_file.seek(0)
                    df1_original = pd.read_excel(banco_file, engine='openpyxl')
                    interno_file.seek(0)
                    df2_original = pd.read_excel(interno_file, engine='openpyxl')
                    
                    # Reconciliar
                    results = reconcile_data(data1, data2, "Banco", "Interno")
                    
                    # Calcular estadísticas
                    conciliados = [r for r in results if r['estado'] == 'Conciliado']
                    exactos = [r for r in conciliados if r['origen'] == 'exacto_fecha_pv_monto']
                    fecha_pv = [r for r in conciliados if r['origen'] == 'fecha_pv']
                    fecha_monto = [r for r in conciliados if r['origen'] == 'fecha_monto']
                    
                    # Guardar resultados
                    st.session_state['results'] = results
                    st.session_state['sheet_names'] = ["Banco", "Interno"]
                    st.session_state['original_files_data'] = {
                        'original_df1': df1_original,
                        'original_df2': df2_original
                    }
                    
                    st.success(f"✅ Conciliación completada!")
                    st.info(f"""
                    **Resultados del cruce:**
                    - Total registros procesados: Banco: {len(data1)}, Interno: {len(data2)}
                    - **Conciliados: {len(conciliados)}**
                      - Exactos (Fecha+PV+Monto): {len(exactos)}
                      - Por Fecha+PV: {len(fecha_pv)}
                      - Por Fecha+Monto: {len(fecha_monto)}
                    - No conciliados: {len(results) - len(conciliados)}
                    """)
                    # Evitar st.rerun() que puede causar errores de DOM
                    # Los resultados se mostrarán automáticamente en la siguiente sección
    
    # Mostrar resultados
    if 'results' in st.session_state and st.session_state['results']:
        results = st.session_state['results']
        sheet_names = st.session_state.get('sheet_names', ['Hoja 1', 'Hoja 2'])
        
        # Estadísticas
        conciliados = [r for r in results if r['estado'] == 'Conciliado']
        no_conciliados = [r for r in results if r['estado'] == 'No conciliado']
        
        monto_conciliado = sum(r['banco'].get('monto', 0) or r['interno'].get('monto', 0) 
                              if r['interno'] else r['banco'].get('monto', 0) 
                              for r in conciliados)
        monto_no_conciliado = sum(r['banco'].get('monto', 0) or r['interno'].get('monto', 0) 
                                 if r['interno'] else r['banco'].get('monto', 0) 
                                 for r in no_conciliados)
        
        st.header("📊 Panel de Control")
        
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.markdown(f"""
            <div class="stat-card info">
                <div class="stat-value">{len(results)}</div>
                <div class="stat-label">Total Registros</div>
            </div>
            """, unsafe_allow_html=True)
        with col2:
            st.markdown(f"""
            <div class="stat-card success">
                <div class="stat-value">{len(conciliados)}</div>
                <div class="stat-label">Conciliados</div>
            </div>
            """, unsafe_allow_html=True)
        with col3:
            st.markdown(f"""
            <div class="stat-card warning">
                <div class="stat-value">{len(no_conciliados)}</div>
                <div class="stat-label">No Conciliados</div>
            </div>
            """, unsafe_allow_html=True)
        with col4:
            st.markdown(f"""
            <div class="stat-card info">
                <div class="stat-value">${monto_conciliado:,.2f}</div>
                <div class="stat-label">Monto Conciliado</div>
            </div>
            """, unsafe_allow_html=True)
        
        # Tabla de resultados
        st.header("📋 Resultados de Conciliación")
        
        # Preparar datos para tabla mostrando claramente qué cruza con qué
        table_data = []
        for r in results:
            banco = r['banco']
            interno = r.get('interno')
            
            # Determinar tipo de coincidencia y mensaje claro
            tipo_coincidencia = ""
            mensaje_cruce = ""
            if r['estado'] == 'Conciliado':
                fila_banco = r.get('fila_banco', '?')
                fila_interno = r.get('fila_interno', '?')
                
                if r['origen'] == 'exacto_fecha_pv_monto':
                    tipo_coincidencia = "✅ EXACTA"
                    mensaje_cruce = f"Fila {fila_banco} ↔ Fila {fila_interno}: Fecha+PV+Monto iguales"
                elif r['origen'] == 'fecha_pv':
                    tipo_coincidencia = "🔄 FECHA+PV"
                    mensaje_cruce = f"Fila {fila_banco} ↔ Fila {fila_interno}: Fecha y PV iguales"
                elif r['origen'] == 'fecha_monto':
                    tipo_coincidencia = "📅 FECHA+MONTO"
                    mensaje_cruce = f"Fila {fila_banco} ↔ Fila {fila_interno}: Fecha y Monto iguales"
            else:
                tipo_coincidencia = "❌ NO CRUZADO"
                if interno:
                    mensaje_cruce = f"Fila {r.get('fila_interno', '?')} de {sheet_names[1]}: No tiene coincidencia"
                else:
                    mensaje_cruce = f"Fila {r.get('fila_banco', '?')} de {sheet_names[0]}: No tiene coincidencia"
            
            table_data.append({
                'Estado': r['estado'],
                'Tipo': tipo_coincidencia,
                'Cruza': mensaje_cruce,
                f'Fila {sheet_names[0]}': r.get('fila_banco', '-'),
                f'Fecha {sheet_names[0]}': banco.get('fecha', '-'),
                f'PV {sheet_names[0]}': banco.get('referencia', '-'),
                f'Monto {sheet_names[0]}': f"${banco.get('monto', 0):,.2f}" if banco.get('monto') else '-',
                f'Fila {sheet_names[1]}': r.get('fila_interno', '-'),
                f'Fecha {sheet_names[1]}': interno.get('fecha', '-') if interno else '-',
                f'PV {sheet_names[1]}': interno.get('referencia', '-') if interno else '-',
                f'Monto {sheet_names[1]}': f"${interno.get('monto', 0):,.2f}" if interno and interno.get('monto') else '-',
                'Diferencia Monto': f"${abs((banco.get('monto', 0) or 0) - (interno.get('monto', 0) or 0) if interno else 0):,.2f}"
            })
        
        df_results = pd.DataFrame(table_data)
        
        # Mostrar tabla con estilo
        st.dataframe(
            df_results,
            use_container_width=True,
            height=400,
            hide_index=True
        )
        
        # Mostrar resumen de cruces
        st.subheader("📊 Resumen de Cruces")
        conciliados_list = [r for r in results if r['estado'] == 'Conciliado']
        
        if conciliados_list:
            st.success(f"✅ Se encontraron {len(conciliados_list)} valores que coinciden entre ambas hojas:")
            
            # Mostrar ejemplos de cruces
            with st.expander("🔍 Ver detalles de los cruces", expanded=True):
                for i, r in enumerate(conciliados_list[:10], 1):  # Mostrar primeros 10
                    banco = r['banco']
                    interno = r['interno']
                    fila_b = r.get('fila_banco', '?')
                    fila_i = r.get('fila_interno', '?')
                    
                    st.markdown(f"""
                    **Cruze #{i}:**
                    - **{sheet_names[0]} (Fila {fila_b})**: Fecha={banco.get('fecha')}, PV={banco.get('referencia')}, Monto=${banco.get('monto', 0):,.2f}
                    - **{sheet_names[1]} (Fila {fila_i})**: Fecha={interno.get('fecha')}, PV={interno.get('referencia')}, Monto=${interno.get('monto', 0):,.2f}
                    - **Tipo**: {r['origen']}
                    ---
                    """)
                
                if len(conciliados_list) > 10:
                    st.info(f"... y {len(conciliados_list) - 10} cruces más (ver tabla completa abajo)")
        
        # Información adicional sobre el cruce
        with st.expander("ℹ️ ¿Cómo funciona el cruce automático?", expanded=False):
            st.markdown("""
            **La aplicación busca automáticamente valores iguales en ambas hojas:**
            
            1. **PRIORIDAD: Fecha + PV** 🔄
               - Busca registros donde la **Fecha** y el **PV** (Punto de Venta) sean iguales
               - Si además el monto es igual → Coincidencia EXACTA ✅
               - Si el monto es diferente → Coincidencia por Fecha+PV 🔄
            
            2. **FECHA + MONTO** 📅
               - Si no hay coincidencia por Fecha+PV, busca donde coinciden Fecha y Monto
               - (El PV puede ser diferente)
            
            3. **NO CRUZADO** ❌
               - Registros que no encontraron coincidencia en la otra hoja
            
            **La tabla muestra claramente:**
            - Qué fila de cada hoja se cruzó con qué fila
            - Los valores que coincidieron (Fecha, PV, Monto)
            - El tipo de coincidencia encontrada
            """)
        
        # Botón de descarga con formato preservado
        def create_excel_with_format(results, sheet_names):
            """Crea Excel preservando formatos originales"""
            wb = Workbook()
            ws = wb.active
            ws.title = "Conciliación"
            
            # Estilos
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center')
            
            # Encabezados
            headers = [
                'Estado', 'Tipo', 'Cruza',
                f'Fila {sheet_names[0]}', f'Fecha {sheet_names[0]}', f'PV {sheet_names[0]}', f'Monto {sheet_names[0]}',
                f'Fila {sheet_names[1]}', f'Fecha {sheet_names[1]}', f'PV {sheet_names[1]}', f'Monto {sheet_names[1]}',
                'Diferencia Monto'
            ]
            
            # Escribir encabezados con formato
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
            
            # Escribir datos preservando formatos
            for row_idx, r in enumerate(results, 2):
                banco = r['banco']
                interno = r.get('interno')
                
                # Determinar tipo de coincidencia
                tipo_coincidencia = ""
                mensaje_cruce = ""
                if r['estado'] == 'Conciliado':
                    fila_banco = r.get('fila_banco', '?')
                    fila_interno = r.get('fila_interno', '?')
                    
                    if r['origen'] == 'exacto_fecha_pv_monto':
                        tipo_coincidencia = "✅ EXACTA"
                        mensaje_cruce = f"Fila {fila_banco} ↔ Fila {fila_interno}: Fecha+PV+Monto iguales"
                    elif r['origen'] == 'fecha_pv':
                        tipo_coincidencia = "🔄 FECHA+PV"
                        mensaje_cruce = f"Fila {fila_banco} ↔ Fila {fila_interno}: Fecha y PV iguales"
                    elif r['origen'] == 'fecha_monto':
                        tipo_coincidencia = "📅 FECHA+MONTO"
                        mensaje_cruce = f"Fila {fila_banco} ↔ Fila {fila_interno}: Fecha y Monto iguales"
                else:
                    tipo_coincidencia = "❌ NO CRUZADO"
                    if interno:
                        mensaje_cruce = f"Fila {r.get('fila_interno', '?')} de {sheet_names[1]}: No tiene coincidencia"
                    else:
                        mensaje_cruce = f"Fila {r.get('fila_banco', '?')} de {sheet_names[0]}: No tiene coincidencia"
                
                # Preparar valores
                row_data = [
                    r['estado'],
                    tipo_coincidencia,
                    mensaje_cruce,
                    r.get('fila_banco', '-'),
                    banco.get('fecha', '-'),
                    banco.get('referencia', '-'),
                    banco.get('monto', 0) if banco.get('monto') else None,
                    r.get('fila_interno', '-'),
                    interno.get('fecha', '-') if interno else '-',
                    interno.get('referencia', '-') if interno else '-',
                    interno.get('monto', 0) if interno and interno.get('monto') else None,
                    abs((banco.get('monto', 0) or 0) - (interno.get('monto', 0) or 0) if interno else 0)
                ]
                
                # Escribir valores con formato apropiado
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    
                    # Formato para estado
                    if col_idx == 1:  # Estado
                        if value == 'Conciliado':
                            cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    
                    # Formato para montos (columnas 7, 11, 12)
                    if col_idx in [7, 11, 12]:
                        if value is not None and isinstance(value, (int, float)) and value != 0:
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    # Formato para fechas (columnas 5, 9)
                    if col_idx in [5, 9]:
                        if value and value != '-':
                            try:
                                # Intentar convertir a fecha
                                date_val = pd.to_datetime(value)
                                cell.value = date_val
                                cell.number_format = 'dd/mm/yyyy'
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                            except:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Formato para números de fila (columnas 4, 8)
                    if col_idx in [4, 8]:
                        if value != '-':
                            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Ajustar ancho de columnas
            column_widths = {
                'A': 12,  # Estado
                'B': 15,  # Tipo
                'C': 50,  # Cruza
                'D': 12,  # Fila 1
                'E': 12,  # Fecha 1
                'F': 12,  # PV 1
                'G': 15,  # Monto 1
                'H': 12,  # Fila 2
                'I': 12,  # Fecha 2
                'J': 12,  # PV 2
                'K': 15,  # Monto 2
                'L': 15   # Diferencia
            }
            
            for col_letter, width in column_widths.items():
                ws.column_dimensions[col_letter].width = width
            
            # Congelar primera fila
            ws.freeze_panes = 'A2'
            
            return wb
        
        # Crear Excel con formato preservado
        try:
            output = BytesIO()
            wb = create_excel_with_format(results, sheet_names)
            
            # Agregar hojas originales preservando formato y valores originales
            if 'original_files_data' in st.session_state:
                try:
                    original_data = st.session_state['original_files_data']
                    
                    # Crear conjuntos de filas que cruzaron (fila_banco y fila_interno son base 1)
                    filas_banco_cruzadas = set()
                    filas_interno_cruzadas = set()
                    for r in results:
                        if r['estado'] == 'Conciliado':
                            if r.get('fila_banco'):
                                filas_banco_cruzadas.add(r['fila_banco'])
                            if r.get('fila_interno'):
                                filas_interno_cruzadas.add(r['fila_interno'])
                    
                    # Definir border para las hojas originales
                    border_original = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Agregar hoja con datos originales de la primera hoja
                    if 'original_df1' in original_data:
                        df1 = original_data['original_df1']
                        # Limpiar nombre de hoja (máximo 31 caracteres, sin caracteres inválidos)
                        sheet_name1 = f"Original_{sheet_names[0]}"[:31].replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_')
                        ws1 = wb.create_sheet(title=sheet_name1)
                        
                        # Escribir encabezados con formato (incluyendo columna Cruce)
                        num_cols = len(df1.columns)
                        for c_idx, col_name in enumerate(df1.columns, 1):
                            cell = ws1.cell(row=1, column=c_idx, value=col_name)
                            cell.font = Font(bold=True, size=11)
                            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                            cell.border = border_original
                        
                        # Agregar encabezado de columna "Cruce"
                        cell_cruce_header = ws1.cell(row=1, column=num_cols + 1, value="Cruce")
                        cell_cruce_header.font = Font(bold=True, size=11)
                        cell_cruce_header.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                        cell_cruce_header.border = border_original
                        cell_cruce_header.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Escribir datos preservando tipos y formatos originales
                        for r_idx, (df_idx, row) in enumerate(df1.iterrows(), 2):
                            # Escribir datos originales
                            for c_idx, (col_name, value) in enumerate(zip(df1.columns, row), 1):
                                cell = ws1.cell(row=r_idx, column=c_idx, value=value)
                                cell.border = border_original
                                
                                # Preservar formato de fecha si es fecha
                                if pd.api.types.is_datetime64_any_dtype(df1[col_name]):
                                    try:
                                        cell.number_format = 'dd/mm/yyyy'
                                    except:
                                        pass
                                
                                # Preservar formato numérico
                                if pd.api.types.is_numeric_dtype(df1[col_name]) and not pd.isna(value):
                                    cell.number_format = '#,##0.00'
                                    cell.alignment = Alignment(horizontal='right', vertical='center')
                            
                            # Agregar columna "Cruce"
                            # df_idx es el índice del DataFrame original (0-based)
                            # El número de fila en Excel original es df_idx + 2 (1 para encabezado + 1 porque Excel empieza en 1)
                            fila_num = int(df_idx) + 2 if isinstance(df_idx, (int, float)) else r_idx
                            cruce_value = "OK" if fila_num in filas_banco_cruzadas else "NO CRUZA"
                            cell_cruce = ws1.cell(row=r_idx, column=num_cols + 1, value=cruce_value)
                            cell_cruce.border = border_original
                            cell_cruce.alignment = Alignment(horizontal='center', vertical='center')
                            
                            # Formato de color para la columna Cruce
                            if cruce_value == "OK":
                                cell_cruce.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                cell_cruce.font = Font(bold=True, color="006100")
                            else:
                                cell_cruce.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                cell_cruce.font = Font(bold=True, color="9C0006")
                        
                        # Ajustar ancho de columnas
                        for c_idx, col_name in enumerate(df1.columns, 1):
                            col_letter = get_column_letter(c_idx)
                            ws1.column_dimensions[col_letter].width = max(len(str(col_name)), 12)
                        # Ajustar ancho de columna Cruce
                        col_letter_cruce = get_column_letter(num_cols + 1)
                        ws1.column_dimensions[col_letter_cruce].width = 12
                    
                    # Agregar hoja con datos originales de la segunda hoja
                    if 'original_df2' in original_data:
                        df2 = original_data['original_df2']
                        # Limpiar nombre de hoja (máximo 31 caracteres, sin caracteres inválidos)
                        sheet_name2 = f"Original_{sheet_names[1]}"[:31].replace('/', '_').replace('\\', '_').replace('?', '_').replace('*', '_').replace('[', '_').replace(']', '_')
                        ws2 = wb.create_sheet(title=sheet_name2)
                        
                        # Escribir encabezados con formato (incluyendo columna Cruce)
                        num_cols2 = len(df2.columns)
                        for c_idx, col_name in enumerate(df2.columns, 1):
                            cell = ws2.cell(row=1, column=c_idx, value=col_name)
                            cell.font = Font(bold=True, size=11)
                            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                            cell.border = border_original
                        
                        # Agregar encabezado de columna "Cruce"
                        cell_cruce_header2 = ws2.cell(row=1, column=num_cols2 + 1, value="Cruce")
                        cell_cruce_header2.font = Font(bold=True, size=11)
                        cell_cruce_header2.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                        cell_cruce_header2.border = border_original
                        cell_cruce_header2.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Escribir datos preservando tipos y formatos originales
                        for r_idx, (df_idx, row) in enumerate(df2.iterrows(), 2):
                            # Escribir datos originales
                            for c_idx, (col_name, value) in enumerate(zip(df2.columns, row), 1):
                                cell = ws2.cell(row=r_idx, column=c_idx, value=value)
                                cell.border = border_original
                                
                                # Preservar formato de fecha si es fecha
                                if pd.api.types.is_datetime64_any_dtype(df2[col_name]):
                                    try:
                                        cell.number_format = 'dd/mm/yyyy'
                                    except:
                                        pass
                                
                                # Preservar formato numérico
                                if pd.api.types.is_numeric_dtype(df2[col_name]) and not pd.isna(value):
                                    cell.number_format = '#,##0.00'
                                    cell.alignment = Alignment(horizontal='right', vertical='center')
                            
                            # Agregar columna "Cruce"
                            # df_idx es el índice del DataFrame original (0-based)
                            # El número de fila en Excel original es df_idx + 2 (1 para encabezado + 1 porque Excel empieza en 1)
                            fila_num = int(df_idx) + 2 if isinstance(df_idx, (int, float)) else r_idx
                            cruce_value = "OK" if fila_num in filas_interno_cruzadas else "NO CRUZA"
                            cell_cruce = ws2.cell(row=r_idx, column=num_cols2 + 1, value=cruce_value)
                            cell_cruce.border = border_original
                            cell_cruce.alignment = Alignment(horizontal='center', vertical='center')
                            
                            # Formato de color para la columna Cruce
                            if cruce_value == "OK":
                                cell_cruce.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                                cell_cruce.font = Font(bold=True, color="006100")
                            else:
                                cell_cruce.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                                cell_cruce.font = Font(bold=True, color="9C0006")
                        
                        # Ajustar ancho de columnas
                        for c_idx, col_name in enumerate(df2.columns, 1):
                            col_letter = get_column_letter(c_idx)
                            ws2.column_dimensions[col_letter].width = max(len(str(col_name)), 12)
                        # Ajustar ancho de columna Cruce
                        col_letter_cruce2 = get_column_letter(num_cols2 + 1)
                        ws2.column_dimensions[col_letter_cruce2].width = 12
                except Exception as e:
                    st.warning(f"Nota: No se pudieron incluir las hojas originales: {str(e)}")
            
            wb.save(output)
            output.seek(0)
            
            st.download_button(
                label="📥 Descargar Excel Conciliado (con formato preservado)",
                data=output.getvalue(),
                file_name=f"conciliacion_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        except Exception as e:
            st.error(f"❌ Error al generar el archivo Excel: {str(e)}")
            st.exception(e)
        
        # Botón para limpiar
        if st.button("🔄 Nueva Conciliación"):
            if 'results' in st.session_state:
                del st.session_state['results']
            st.rerun()

if __name__ == "__main__":
    main()

